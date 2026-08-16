"""Ledgerly Personal - user_id-scoped endpoints, kept out of server.py per
design/ledgerly-personal.md. No business_id anywhere in this file - that's
the structural guarantee against Personal/Business data mixing, not just a
filter convention.
"""
import asyncio
import calendar
import csv
import io
import json
import uuid
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Literal, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from groq import AsyncGroq, AuthenticationError as GroqAuthenticationError, RateLimitError as GroqRateLimitError, APIStatusError as GroqAPIStatusError
from pydantic import BaseModel
from reportlab.lib.pagesizes import LETTER
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from server import (
    db, get_current_user, now_utc, VAPID_PUBLIC_KEY, VAPID_PRIVATE_KEY, _send_one_push,
    _resolve_ai_key, _weeks_in_month, _fmt, _fmt_date, _pnl_rows, _pnl_pdf_story, _pdf_ledgerly_header,
)

personal_router = APIRouter(prefix="/api/personal")

DUE_SOON_DAYS = 3  # matches BillsScreen.jsx's own client-side threshold


class PersonalTransactionIn(BaseModel):
    type: Literal["income", "expense"]
    amount: float
    category: str
    description: Optional[str] = ""
    date: str  # ISO date
    currency: str = "USD"
    bill_id: Optional[str] = None
    receipt_image: Optional[str] = None  # base64, from POST /receipts/extract
    receipt_content_type: Optional[str] = None


class PersonalBudgetIn(BaseModel):
    category: str
    monthly_limit: float
    currency: str = "USD"


class PersonalBillIn(BaseModel):
    name: str
    category: str
    amount: float
    due_date: str  # ISO date
    currency: str = "USD"
    recurring: bool = True


class PersonalGoalIn(BaseModel):
    name: str
    target_amount: float
    target_date: Optional[str] = None
    currency: str = "USD"


class PersonalGoalContributionIn(BaseModel):
    amount: float
    date: str  # ISO date
    note: Optional[str] = ""


class PersonalChatIn(BaseModel):
    message: str
    conversation_id: Optional[str] = None


class PersonalConversationRenameIn(BaseModel):
    title: str


# ---- Personal Transactions ----
@personal_router.get("/transactions")
async def list_personal_transactions(
    type: Optional[Literal["income", "expense"]] = None,
    category: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    has_receipt: Optional[bool] = None,
    user=Depends(get_current_user),
):
    query = {"user_id": user["user_id"]}
    if type:
        query["type"] = type
    if category:
        query["category"] = category
    if has_receipt:
        query["receipt_image"] = {"$ne": None}
    if date_from or date_to:
        date_query = {}
        if date_from:
            date_query["$gte"] = date_from
        if date_to:
            date_query["$lte"] = date_to
        query["date"] = date_query
    cursor = db.personal_transactions.find(query, {"_id": 0}).sort("date", -1)
    return await cursor.to_list(2000)

@personal_router.post("/transactions")
async def create_personal_transaction(payload: PersonalTransactionIn, user=Depends(get_current_user)):
    tx = payload.model_dump()
    tx["id"] = str(uuid.uuid4())
    tx["user_id"] = user["user_id"]
    tx["created_at"] = now_utc().isoformat()
    await db.personal_transactions.insert_one(tx)
    tx.pop("_id", None)
    if tx.get("bill_id"):
        # This transaction *is* the payment record - advance/clear the bill
        # without inserting a second transaction (unlike POST
        # /bills/{id}/mark-paid, which has no user-entered transaction to
        # reuse and must create one itself).
        bill = await db.personal_bills.find_one({"id": tx["bill_id"], "user_id": user["user_id"]}, {"_id": 0})
        if not bill:
            raise HTTPException(status_code=404, detail="Bill not found")
        await _advance_or_clear_bill(bill)
    return tx

@personal_router.put("/transactions/{tx_id}")
async def update_personal_transaction(tx_id: str, payload: PersonalTransactionIn, user=Depends(get_current_user)):
    upd = payload.model_dump()
    res = await db.personal_transactions.update_one({"id": tx_id, "user_id": user["user_id"]}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return await db.personal_transactions.find_one({"id": tx_id}, {"_id": 0})

@personal_router.delete("/transactions/{tx_id}")
async def delete_personal_transaction(tx_id: str, user=Depends(get_current_user)):
    res = await db.personal_transactions.delete_one({"id": tx_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ---- Dashboard series (mirrors server.dashboard_series, minus invoices -
# personal has none) ----
@personal_router.get("/reports/series")
async def personal_dashboard_series(
    granularity: Literal["day", "week", "month", "year"] = Query("month"),
    date: Optional[str] = Query(None, description="Exact day, for granularity=day"),
    year: Optional[int] = Query(None, description="For week/month/year"),
    month: Optional[int] = Query(None, ge=1, le=12, description="For week/month"),
    week: Optional[int] = Query(None, ge=1, le=6, description="Which week of the month, for granularity=week"),
    user=Depends(get_current_user),
):
    today = now_utc()

    if granularity == "day":
        try:
            start = datetime.fromisoformat(date) if date else today
        except Exception:
            start = today
        start = start.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)
        end = start
        buckets = [start.date().isoformat()]

        def bucket_key(d: datetime) -> str:
            return d.date().isoformat()

        label = f"{calendar.month_abbr[start.month]} {start.day}, {start.year}"
    elif granularity == "week":
        y = year or today.year
        m = month or today.month
        days_in_month = calendar.monthrange(y, m)[1]
        w = max(1, min(week or 1, _weeks_in_month(y, m)))
        start_day = (w - 1) * 7 + 1
        end_day = min(w * 7, days_in_month)
        start = datetime(y, m, start_day)
        end = datetime(y, m, end_day)
        buckets = [(start + timedelta(days=i)).date().isoformat() for i in range((end - start).days + 1)]

        def bucket_key(d: datetime) -> str:
            return d.date().isoformat()

        label = f"Week {w} of {calendar.month_name[m]} {y}"
    elif granularity == "month":
        y = year or today.year
        m = month or today.month
        start = datetime(y, m, 1)
        days_in_month = calendar.monthrange(y, m)[1]
        end = datetime(y, m, days_in_month)
        num_weeks = _weeks_in_month(y, m)
        buckets = [f"{y}-{m:02d}-W{w}" for w in range(1, num_weeks + 1)]

        def bucket_key(d: datetime) -> str:
            return f"{d.year}-{d.month:02d}-W{(d.day - 1) // 7 + 1}"

        label = f"{calendar.month_name[m]} {y}"
    else:  # year
        y = year or today.year
        start = datetime(y, 1, 1)
        end = datetime(y, 12, 31)
        buckets = [f"{y}-{mm:02d}" for mm in range(1, 13)]

        def bucket_key(d: datetime) -> str:
            return f"{d.year}-{d.month:02d}"

        label = str(y)

    start_s, end_s = start.date().isoformat(), end.date().isoformat()

    def bucket_label(b: str) -> str:
        if granularity == "day":
            return label
        if granularity == "week":
            d = datetime.fromisoformat(b)
            return calendar.day_abbr[d.weekday()]
        if granularity == "month":
            return f"Week {b.rsplit('W', 1)[1]}"
        mm = int(b.split("-")[1])
        return calendar.month_abbr[mm]

    series = {b: {"income": 0.0, "expense": 0.0} for b in buckets}

    txs = await db.personal_transactions.find(
        {"user_id": user["user_id"], "date": {"$gte": start_s, "$lte": end_s}}, {"_id": 0}
    ).to_list(20000)
    cats_income, cats_expense = defaultdict(float), defaultdict(float)
    total_income = total_expense = 0.0
    for t in txs:
        try:
            d = datetime.fromisoformat(t["date"])
        except Exception:
            continue
        key = bucket_key(d)
        if key not in series:
            continue
        amt = t["amount"]
        if t["type"] == "income":
            series[key]["income"] += amt
            total_income += amt
            cats_income[t["category"]] += amt
        else:
            series[key]["expense"] += amt
            total_expense += amt
            cats_expense[t["category"]] += amt

    return {
        "window": {"start": start_s, "end": end_s, "granularity": granularity, "label": label},
        "series": [
            {
                "period": b,
                "label": bucket_label(b),
                "income": round(series[b]["income"], 2),
                "expense": round(series[b]["expense"], 2),
                "net": round(series[b]["income"] - series[b]["expense"], 2),
            }
            for b in buckets
        ],
        "categories": {
            "income": [{"name": k, "value": round(v, 2)} for k, v in cats_income.items()],
            "expense": [{"name": k, "value": round(v, 2)} for k, v in cats_expense.items()],
        },
        "totals": {
            "income": round(total_income, 2),
            "expenses": round(total_expense, 2),
            "net": round(total_income - total_expense, 2),
        },
        "transactions_count": len(txs),
    }


# ---- Reports (income & expenses only - personal_transactions has no
# tax_amount field, so there's no personal equivalent of business's
# separate /reports/tax + tax export) ----
async def _personal_pnl(start: str, end: str, user: dict) -> dict:
    txs = await db.personal_transactions.find(
        {"user_id": user["user_id"], "date": {"$gte": start, "$lte": end}}, {"_id": 0}
    ).to_list(20000)
    income_by_cat, expense_by_cat = {}, {}
    for t in txs:
        bucket = income_by_cat if t["type"] == "income" else expense_by_cat
        bucket[t["category"]] = bucket.get(t["category"], 0) + t["amount"]
    total_income = sum(income_by_cat.values())
    total_expense = sum(expense_by_cat.values())
    return {
        "start": start, "end": end,
        "income": [{"category": k, "amount": round(v, 2)} for k, v in income_by_cat.items()],
        "expenses": [{"category": k, "amount": round(v, 2)} for k, v in expense_by_cat.items()],
        "total_income": round(total_income, 2),
        "total_expense": round(total_expense, 2),
        "net": round(total_income - total_expense, 2),
    }

@personal_router.get("/reports/pnl")
async def personal_pnl_report(start: str = Query(...), end: str = Query(...), user=Depends(get_current_user)):
    return await _personal_pnl(start, end, user)

@personal_router.get("/export/pnl")
async def export_personal_pnl(
    format: str = Query("csv"), start: str = Query(...), end: str = Query(...), user=Depends(get_current_user),
):
    data = await _personal_pnl(start, end, user)
    rows_iter = list(_pnl_rows(data))

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in rows_iter:
            writer.writerow(row)
        content = buf.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content), media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="income_expenses.csv"'},
        )
    elif format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income & Expenses"
        for row in rows_iter:
            ws.append(row)
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="income_expenses.xlsx"'},
        )
    elif format == "pdf":
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=LETTER, leftMargin=0.6 * inch, rightMargin=0.6 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        story = _pdf_ledgerly_header(styles)
        label_style = ParagraphStyle('rlabel', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor("#64748b"))
        range_style = ParagraphStyle('rtitle', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor("#0f172a"), spaceBefore=2)
        story.append(Paragraph(f"<b>INCOME &amp; EXPENSES</b> &middot; {user.get('name')}", label_style))
        story.append(Paragraph(f"{_fmt_date(start)} to {_fmt_date(end)}", range_style))
        story.append(Spacer(1, 10))
        story.extend(_pnl_pdf_story(data, user.get("currency", "USD")))
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": 'attachment; filename="income_expenses.pdf"'},
        )
    else:
        raise HTTPException(status_code=400, detail="Unknown format")

def _months_in_range(start: str, end: str) -> int:
    """Same accounting as the frontend's monthsInRange - a budget's
    monthly_limit is a single ongoing value, not versioned per month, so
    comparing it against a multi-month range means scaling by how many
    calendar months the range touches."""
    s = datetime.fromisoformat(start)
    e = datetime.fromisoformat(end)
    if e < s:
        return 0
    return max((e.year - s.year) * 12 + (e.month - s.month) + 1, 0)

def _category_export_rows(category: str, txs: list, budget_info: Optional[dict]):
    if budget_info:
        yield ["Metric", "Value"]
        yield ["Category", category]
        yield ["Monthly limit", budget_info["monthly_limit"]]
        yield ["Months in period", budget_info["months"]]
        yield ["Limit over period", budget_info["period_limit"]]
        yield ["Total spent", budget_info["total"]]
        over = budget_info["total"] > budget_info["period_limit"]
        yield ["Over by" if over else "Under by", round(abs(budget_info["period_limit"] - budget_info["total"]), 2)]
        yield []
    yield ["Date", "Type", "Description", "Amount", "Currency"]
    for t in txs:
        yield [_fmt_date(t["date"]), t["type"], t.get("description", ""), t["amount"], t.get("currency", "USD")]

@personal_router.get("/export/category")
async def export_personal_category(
    format: str = Query("csv"),
    category: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    budget_id: Optional[str] = Query(None),
    user=Depends(get_current_user),
):
    txs = await db.personal_transactions.find(
        {"user_id": user["user_id"], "category": category, "date": {"$gte": start, "$lte": end}}, {"_id": 0}
    ).sort("date", -1).to_list(20000)
    total = round(sum(t["amount"] for t in txs), 2)

    budget_info = None
    if budget_id:
        budget = await db.personal_budgets.find_one({"id": budget_id, "user_id": user["user_id"]}, {"_id": 0})
        if budget:
            months = _months_in_range(start, end)
            budget_info = {
                "monthly_limit": budget["monthly_limit"],
                "months": months,
                "period_limit": round(budget["monthly_limit"] * months, 2),
                "total": total,
            }

    rows_iter = list(_category_export_rows(category, txs, budget_info))
    safe_name = "".join(c if c.isalnum() else "_" for c in category.lower())

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in rows_iter:
            writer.writerow(row)
        content = buf.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(content), media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_report.csv"'},
        )
    elif format == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = category[:31] or "Report"
        for row in rows_iter:
            ws.append(row)
        header_row = 9 if budget_info else 1
        for cell in ws[header_row]:
            cell.font = openpyxl.styles.Font(bold=True)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_report.xlsx"'},
        )
    elif format == "pdf":
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=LETTER, leftMargin=0.6 * inch, rightMargin=0.6 * inch, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        styles = getSampleStyleSheet()
        story = _pdf_ledgerly_header(styles)
        label_style = ParagraphStyle('rlabel', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor("#64748b"))
        range_style = ParagraphStyle('rtitle', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor("#0f172a"), spaceBefore=2)
        story.append(Paragraph(f"<b>{'BUDGET' if budget_info else 'CATEGORY'} REPORT</b> &middot; {user.get('name')}", label_style))
        story.append(Paragraph(category, range_style))
        story.append(Paragraph(f"{_fmt_date(start)} to {_fmt_date(end)}", label_style))
        story.append(Spacer(1, 10))

        if budget_info:
            over = budget_info["total"] > budget_info["period_limit"]
            metrics = [
                ["Monthly limit", _fmt(budget_info["monthly_limit"], user.get("currency", "USD"))],
                ["Limit over period", _fmt(budget_info["period_limit"], user.get("currency", "USD"))],
                ["Total spent", _fmt(budget_info["total"], user.get("currency", "USD"))],
                ["Over by" if over else "Under by", _fmt(abs(budget_info["period_limit"] - budget_info["total"]), user.get("currency", "USD"))],
            ]
            mtbl = Table(metrics, colWidths=[2.8 * inch, 2.8 * inch])
            mtbl.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor("#0f172a")),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
            ]))
            story.append(mtbl)
            story.append(Spacer(1, 14))

        tx_header = ["Date", "Type", "Description", "Amount", "Currency"]
        tx_rows = [tx_header] + [
            [_fmt_date(t["date"]), t["type"], t.get("description", ""), _fmt(t["amount"], t.get("currency", user.get("currency", "USD"))), t.get("currency", "USD")]
            for t in txs
        ]
        ttbl = Table(tx_rows, repeatRows=1)
        ttbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#0f172a")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(ttbl)

        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_report.pdf"'},
        )
    else:
        raise HTTPException(status_code=400, detail="Unknown format")


# ---- Budgets ----
# Spend-vs-limit is computed live from personal_transactions matching the
# budget's category (see get_budgets_summary below) - that's the read side
# of the two-way relationship. create/update here is the write side: users
# set the limit, transactions in that category drive what fills it.
@personal_router.post("/budgets")
async def create_personal_budget(payload: PersonalBudgetIn, user=Depends(get_current_user)):
    existing = await db.personal_budgets.find_one({"user_id": user["user_id"], "category": payload.category})
    if existing:
        raise HTTPException(status_code=400, detail=f"A budget for {payload.category} already exists")
    budget = payload.model_dump()
    budget["id"] = str(uuid.uuid4())
    budget["user_id"] = user["user_id"]
    await db.personal_budgets.insert_one(budget)
    budget.pop("_id", None)
    return budget

@personal_router.put("/budgets/{budget_id}")
async def update_personal_budget(budget_id: str, payload: PersonalBudgetIn, user=Depends(get_current_user)):
    res = await db.personal_budgets.update_one(
        {"id": budget_id, "user_id": user["user_id"]}, {"$set": payload.model_dump()}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return await db.personal_budgets.find_one({"id": budget_id}, {"_id": 0})

@personal_router.get("/budgets/summary")
async def get_budgets_summary(month: Optional[str] = None, user=Depends(get_current_user)):
    month = month or now_utc().date().isoformat()[:7]
    budgets = await db.personal_budgets.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(200)
    result = []
    for b in budgets:
        txs = await db.personal_transactions.find(
            {
                "user_id": user["user_id"],
                "category": b["category"],
                "type": "expense",
                "date": {"$gte": f"{month}-01", "$lt": f"{month}-32"},
            },
            {"amount": 1},
        ).to_list(2000)
        spent = sum(t["amount"] for t in txs)
        result.append({
            "id": b["id"],
            "category": b["category"],
            "monthly_limit": b["monthly_limit"],
            "currency": b.get("currency", "USD"),
            "spent": spent,
            "remaining": b["monthly_limit"] - spent,
        })
    return result

@personal_router.delete("/budgets/{budget_id}")
async def delete_personal_budget(budget_id: str, user=Depends(get_current_user)):
    res = await db.personal_budgets.delete_one({"id": budget_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}


# ---- Bills ----
def _advance_one_month(date_str: str) -> str:
    from datetime import date as _date
    d = _date.fromisoformat(date_str)
    month = d.month + 1
    year = d.year + (month - 1) // 12
    month = (month - 1) % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return d.replace(year=year, month=month, day=day).isoformat()

async def _advance_or_clear_bill(bill: dict) -> None:
    if bill.get("recurring"):
        await db.personal_bills.update_one(
            {"id": bill["id"]},
            {"$set": {"due_date": _advance_one_month(bill["due_date"]), "last_reminder_sent_at": None}},
        )
    else:
        await db.personal_bills.delete_one({"id": bill["id"]})

@personal_router.get("/bills")
async def list_personal_bills(user=Depends(get_current_user)):
    cursor = db.personal_bills.find({"user_id": user["user_id"]}, {"_id": 0}).sort("due_date", 1)
    return await cursor.to_list(500)

@personal_router.post("/bills")
async def create_personal_bill(payload: PersonalBillIn, user=Depends(get_current_user)):
    bill = payload.model_dump()
    bill["id"] = str(uuid.uuid4())
    bill["user_id"] = user["user_id"]
    bill["last_reminder_sent_at"] = None
    await db.personal_bills.insert_one(bill)
    bill.pop("_id", None)
    return bill

@personal_router.put("/bills/{bill_id}")
async def update_personal_bill(bill_id: str, payload: PersonalBillIn, user=Depends(get_current_user)):
    upd = payload.model_dump()
    # due_date may have moved - clear the one-shot reminder guard so
    # _check_bills_due_soon re-evaluates against the new date next cycle.
    upd["last_reminder_sent_at"] = None
    res = await db.personal_bills.update_one({"id": bill_id, "user_id": user["user_id"]}, {"$set": upd})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return await db.personal_bills.find_one({"id": bill_id}, {"_id": 0})

@personal_router.delete("/bills/{bill_id}")
async def delete_personal_bill(bill_id: str, user=Depends(get_current_user)):
    res = await db.personal_bills.delete_one({"id": bill_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}

@personal_router.post("/bills/{bill_id}/mark-paid")
async def mark_bill_paid(bill_id: str, user=Depends(get_current_user)):
    bill = await db.personal_bills.find_one({"id": bill_id, "user_id": user["user_id"]}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Not found")
    tx = {
        "id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        "type": "expense",
        "amount": bill["amount"],
        "category": bill["category"],
        "description": bill["name"],
        "date": now_utc().date().isoformat(),
        "currency": bill.get("currency", "USD"),
        "bill_id": bill["id"],
        "created_at": now_utc().isoformat(),
    }
    await db.personal_transactions.insert_one(tx)
    tx.pop("_id", None)
    await _advance_or_clear_bill(bill)
    return tx


# ---- Savings goals ----
@personal_router.get("/goals")
async def list_personal_goals(user=Depends(get_current_user)):
    goals = await db.personal_savings_goals.find({"user_id": user["user_id"]}, {"_id": 0}).to_list(200)
    result = []
    for g in goals:
        contributions = await db.personal_goal_contributions.find(
            {"goal_id": g["id"], "user_id": user["user_id"]}, {"amount": 1}
        ).to_list(2000)
        result.append({**g, "current_amount": sum(c["amount"] for c in contributions)})
    return result

@personal_router.post("/goals")
async def create_personal_goal(payload: PersonalGoalIn, user=Depends(get_current_user)):
    goal = payload.model_dump()
    goal["id"] = str(uuid.uuid4())
    goal["user_id"] = user["user_id"]
    await db.personal_savings_goals.insert_one(goal)
    goal.pop("_id", None)
    return {**goal, "current_amount": 0}

@personal_router.put("/goals/{goal_id}")
async def update_personal_goal(goal_id: str, payload: PersonalGoalIn, user=Depends(get_current_user)):
    res = await db.personal_savings_goals.update_one(
        {"id": goal_id, "user_id": user["user_id"]}, {"$set": payload.model_dump()}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    goal = await db.personal_savings_goals.find_one({"id": goal_id}, {"_id": 0})
    contributions = await db.personal_goal_contributions.find(
        {"goal_id": goal_id, "user_id": user["user_id"]}, {"amount": 1}
    ).to_list(2000)
    return {**goal, "current_amount": sum(c["amount"] for c in contributions)}

@personal_router.post("/goals/{goal_id}/contributions")
async def add_goal_contribution(goal_id: str, payload: PersonalGoalContributionIn, user=Depends(get_current_user)):
    goal = await db.personal_savings_goals.find_one({"id": goal_id, "user_id": user["user_id"]}, {"_id": 0})
    if not goal:
        raise HTTPException(status_code=404, detail="Not found")
    existing = await db.personal_goal_contributions.find(
        {"goal_id": goal_id, "user_id": user["user_id"]}, {"amount": 1}
    ).to_list(2000)
    previous_amount = sum(c["amount"] for c in existing)

    contrib = payload.model_dump()
    contrib["id"] = str(uuid.uuid4())
    contrib["goal_id"] = goal_id
    contrib["user_id"] = user["user_id"]
    await db.personal_goal_contributions.insert_one(contrib)
    contrib.pop("_id", None)

    current_amount = previous_amount + contrib["amount"]
    # "just_reached" only fires on the contribution that crosses the line,
    # not on every contribution made after a goal is already complete.
    just_reached = previous_amount < goal["target_amount"] <= current_amount
    return {"contribution": contrib, "current_amount": current_amount, "just_reached": just_reached}

@personal_router.delete("/goals/{goal_id}")
async def delete_personal_goal(goal_id: str, user=Depends(get_current_user)):
    res = await db.personal_savings_goals.delete_one({"id": goal_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    await db.personal_goal_contributions.delete_many({"goal_id": goal_id, "user_id": user["user_id"]})
    return {"success": True}

@personal_router.get("/goals/{goal_id}/contributions")
async def list_goal_contributions(goal_id: str, user=Depends(get_current_user)):
    cursor = db.personal_goal_contributions.find(
        {"goal_id": goal_id, "user_id": user["user_id"]}, {"_id": 0}
    ).sort("date", -1)
    return await cursor.to_list(2000)


# ---- AI Insights ----
# Reuses server._resolve_ai_key. A personal-only account (no business at all)
# has no business_id, so it rides the shared Groq key with no per-business
# quota cap - _resolve_ai_key handles that case explicitly. An account that
# does have a business shares that business's key/quota regardless of
# active_context. Everything else here stays user_id-scoped like the rest of
# this file.
async def _personal_context(user: dict) -> str:
    user_id = user["user_id"]
    currency = user.get("currency", "USD")
    txs = await db.personal_transactions.find({"user_id": user_id}, {"_id": 0}).to_list(2000)
    income = sum(t["amount"] for t in txs if t["type"] == "income")
    expenses = sum(t["amount"] for t in txs if t["type"] == "expense")
    by_category = {}
    for t in txs:
        if t["type"] == "expense":
            by_category[t["category"]] = by_category.get(t["category"], 0) + t["amount"]

    budgets = await get_budgets_summary(month=None, user=user)
    bills = await db.personal_bills.find({"user_id": user_id}, {"_id": 0}).sort("due_date", 1).to_list(500)
    goals = await list_personal_goals(user=user)

    budgets_str = "; ".join(f"{b['category']}: {b['spent']:.2f}/{b['monthly_limit']:.2f}" for b in budgets) or "none set"
    bills_str = "; ".join(f"{b['name']} ({b['category']}) ${b['amount']:.2f} due {b['due_date']}" for b in bills) or "none"
    goals_str = "; ".join(f"{g['name']}: {g['current_amount']:.2f}/{g['target_amount']:.2f}" for g in goals) or "none"

    return f"""Currency: {currency}

Financial Summary (all time):
- Total Income: {income:.2f}
- Total Expenses: {expenses:.2f}
- Net: {income - expenses:.2f}
- Total Transactions: {len(txs)}

Expense Categories (all time): {by_category}

Budgets this month (category: spent/limit): {budgets_str}
Upcoming Bills: {bills_str}
Savings Goals (current/target): {goals_str}
"""

@personal_router.post("/insights/chat")
async def personal_insights_chat(payload: PersonalChatIn, user=Depends(get_current_user)):
    message = payload.message.strip()
    if not message:
        raise HTTPException(status_code=400, detail="Message cannot be empty")
    api_key = await _resolve_ai_key(user)

    conversation = None
    if payload.conversation_id:
        conversation = await db.personal_ai_conversations.find_one(
            {"conversation_id": payload.conversation_id, "user_id": user["user_id"]}, {"_id": 0}
        )
        if not conversation:
            raise HTTPException(status_code=404, detail="Conversation not found")

    history = conversation["messages"] if conversation else []
    ctx = await _personal_context(user)
    messages = [{
        "role": "system",
        "content": (
            "You are a friendly personal finance coach helping someone understand their own money. "
            "Be direct, concrete, and cite specific numbers. Use plain markdown: paragraphs, **bold**, and bullet "
            "lists. Never use markdown tables (pipe/dash grid syntax) - the chat UI doesn't render them, so present "
            "any comparison or breakdown as a short bullet list or prose instead. Never invent numbers not provided. "
            "Keep answers focused and conversational - structure with headings/bullets only when it genuinely helps.\n\n"
            f"Current personal finance data:\n{ctx}"
        ),
    }]
    messages += [{"role": m["role"], "content": m["content"]} for m in history]
    messages.append({"role": "user", "content": message})

    conversation_id = conversation["conversation_id"] if conversation else str(uuid.uuid4())
    now_iso = now_utc().isoformat()

    def sse(payload: dict) -> str:
        return f"data: {json.dumps(payload)}\n\n"

    async def event_stream():
        yield sse({"type": "meta", "conversation_id": conversation_id})
        reply_parts = []
        client = AsyncGroq(api_key=api_key)
        try:
            stream = await client.chat.completions.create(
                model="openai/gpt-oss-120b",
                messages=messages,
                max_completion_tokens=2048,
                reasoning_effort="low",
                stream=True,
            )
            async for chunk in stream:
                text = chunk.choices[0].delta.content
                if text:
                    reply_parts.append(text)
                    yield sse({"type": "chunk", "text": text})
        except GroqAuthenticationError:
            yield sse({"type": "error", "message": "Invalid Groq API key. Update it in Settings → Business → AI Insights."})
            return
        except GroqRateLimitError:
            yield sse({"type": "error", "message": "Groq API rate limit or quota exceeded. Please try again shortly."})
            return
        except GroqAPIStatusError as e:
            yield sse({"type": "error", "message": f"AI service error: {e.message}"})
            return
        except Exception:
            yield sse({"type": "error", "message": "AI service temporarily unavailable. Please try again shortly."})
            return

        reply = "".join(reply_parts)
        new_messages = [
            {"role": "user", "content": message, "at": now_iso},
            {"role": "assistant", "content": reply, "at": now_iso},
        ]
        if conversation:
            await db.personal_ai_conversations.update_one(
                {"conversation_id": conversation_id},
                {"$push": {"messages": {"$each": new_messages}}, "$set": {"updated_at": now_iso}},
            )
        else:
            title = message if len(message) <= 60 else message[:57] + "..."
            await db.personal_ai_conversations.insert_one({
                "conversation_id": conversation_id,
                "user_id": user["user_id"],
                "title": title,
                "messages": new_messages,
                "created_at": now_iso,
                "updated_at": now_iso,
            })
        yield sse({"type": "done"})

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@personal_router.get("/insights/conversations")
async def list_personal_conversations(user=Depends(get_current_user)):
    return await db.personal_ai_conversations.find(
        {"user_id": user["user_id"]}, {"_id": 0, "conversation_id": 1, "title": 1, "updated_at": 1}
    ).sort("updated_at", -1).to_list(100)

@personal_router.get("/insights/conversations/{conversation_id}")
async def get_personal_conversation(conversation_id: str, user=Depends(get_current_user)):
    convo = await db.personal_ai_conversations.find_one(
        {"conversation_id": conversation_id, "user_id": user["user_id"]}, {"_id": 0}
    )
    if not convo:
        raise HTTPException(status_code=404, detail="Conversation not found")
    return convo

@personal_router.put("/insights/conversations/{conversation_id}")
async def rename_personal_conversation(conversation_id: str, payload: PersonalConversationRenameIn, user=Depends(get_current_user)):
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title cannot be empty")
    res = await db.personal_ai_conversations.update_one(
        {"conversation_id": conversation_id, "user_id": user["user_id"]}, {"$set": {"title": title}}
    )
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Conversation not found")
    return {"success": True}

@personal_router.delete("/insights/conversations/{conversation_id}")
async def delete_personal_conversation(conversation_id: str, user=Depends(get_current_user)):
    res = await db.personal_ai_conversations.delete_one(
        {"conversation_id": conversation_id, "user_id": user["user_id"]}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Conversation not found")
    return {"success": True}


# ---- Notifications + push ----
async def _push_to_personal_user(user_id: str, title: str, message: str, link: Optional[str]):
    if not (VAPID_PUBLIC_KEY and VAPID_PRIVATE_KEY):
        return
    subs = await db.push_subscriptions.find({"user_id": user_id, "app": {"$ne": "go"}}, {"_id": 0}).to_list(50)
    if not subs:
        return
    payload = {"title": title, "message": message, "link": link}
    results = await asyncio.gather(*(asyncio.to_thread(_send_one_push, sub, payload) for sub in subs), return_exceptions=True)
    expired = [sub["endpoint"] for sub, r in zip(subs, results) if r == "expired"]
    if expired:
        await db.push_subscriptions.delete_many({"endpoint": {"$in": expired}})

async def _notify_personal(user_id: str, type_: str, title: str, message: str = "", link: Optional[str] = None):
    await db.personal_notifications.insert_one({
        "id": str(uuid.uuid4()), "user_id": user_id, "type": type_, "title": title,
        "message": message, "link": link, "read": False, "created_at": now_utc().isoformat(),
    })
    await _push_to_personal_user(user_id, title, message, link)

async def _check_bills_due_soon(user_id: str):
    """Runs opportunistically on GET /personal/notifications (no background
    job runner in this deployment - mirrors _check_overdue_invoices,
    server.py:963). Uses personal_bills' last_reminder_sent_at as the
    one-shot-per-cycle guard: mark-paid already clears it when a bill
    advances, so this naturally re-fires next cycle without extra state."""
    today = now_utc().date()
    bills = await db.personal_bills.find({"user_id": user_id, "last_reminder_sent_at": None}, {"_id": 0}).to_list(500)
    for b in bills:
        due = datetime.fromisoformat(b["due_date"]).date()
        if due <= today + timedelta(days=DUE_SOON_DAYS):
            days = (due - today).days
            when = "today" if days == 0 else ("tomorrow" if days == 1 else f"in {days} days")
            await _notify_personal(b["user_id"], "bill_due_soon", f"{b['name']} due {when}", f"${b['amount']:.2f} due {b['due_date']}", link="/bills")
            await db.personal_bills.update_one({"id": b["id"]}, {"$set": {"last_reminder_sent_at": now_utc().isoformat()}})

@personal_router.get("/notifications")
async def list_personal_notifications(user=Depends(get_current_user)):
    await _check_bills_due_soon(user["user_id"])
    items = await db.personal_notifications.find({"user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    unread_count = await db.personal_notifications.count_documents({"user_id": user["user_id"], "read": False})
    return {"items": items, "unread_count": unread_count}

@personal_router.post("/notifications/read-all")
async def mark_all_personal_notifications_read(user=Depends(get_current_user)):
    await db.personal_notifications.update_many({"user_id": user["user_id"], "read": False}, {"$set": {"read": True}})
    return {"success": True}

@personal_router.delete("/notifications")
async def clear_personal_notifications(user=Depends(get_current_user)):
    await db.personal_notifications.delete_many({"user_id": user["user_id"]})
    return {"success": True}

@personal_router.delete("/notifications/{notif_id}")
async def delete_personal_notification(notif_id: str, user=Depends(get_current_user)):
    res = await db.personal_notifications.delete_one({"id": notif_id, "user_id": user["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"success": True}
